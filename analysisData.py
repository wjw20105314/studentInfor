#!/usr/bin/env python
# -*- coding: utf-8 -*-

# coding=UTF-8
import traceback

import openpyxl
import configparser
import time
# 写入Excel表需要使用的库
from openpyxl import Workbook

try:

    config = configparser.ConfigParser()
    config.read('properties.conf', encoding='utf-8-sig')

    file_name = config['fileCofig']['file_name']  # Excel文件名称

    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    i = 1
    #教育信息
    setSQLData1 = []
    row1 = ['序号', '镇', '行政村', '户主姓名', '户主证件号码', '姓名', '性别', '身份证号码', '年龄', '文化程度', '在校生状态', '在校生状况', '学校名称', '成员备注']
    setSQLData1.append(row1)
    #16-16（含）-60周岁（不含）健康男性非在校生为无劳动力或弱劳动力或半劳动力
    setSQLData2 = []
    row2 = ['序号', '镇', '行政村', '户主姓名', '户主证件号码', '姓名', '性别', '身份证号码', '年龄', '在校生状态', '是否健康', '劳动能力', '成员备注']
    setSQLData2.append(row2)
    for i in range(ws.max_row):

        i = i + 1  # 下标指向下一行，第二行为数据行

        cell_ce = ws['CE' + str(i)].value  # 读取在校生状态
        cell_cf = ws['CF' + str(i)].value  # 读取在校生状况
        cell_b = ws['B' + str(i)].value  # 读取镇
        cell_c = ws['C' + str(i)].value  # 行政村
        cell_bq = ws['G' + str(i)].value  # 户主姓名
        cell_bs = ws['H' + str(i)].value  # 户主证件号码
        cell_bp = ws['BP' + str(i)].value  # 姓名
        cell_bq = ws['BQ' + str(i)].value  # 性别
        cell_bs = ws['BS' + str(i)].value  # 身份证号
        cell_bz = ws['BZ' + str(i)].value  # 年龄
        cell_cj = ws['CJ' + str(i)].value  # 学校名称
        cell_du = ws['DU' + str(i)].value  # 成员备注
        cell_ck = ws['CK' + str(i)].value  # 劳动能力
        cell_cd = ws['CD' + str(i)].value  # 读取文化程度
        cell_cp = ws['CP' + str(i)].value  # 是否健康
        # 抓异常数据-在校生状态
        if (cell_cd == '小学' and cell_cf != '小学') or (cell_cd == '初中' and cell_cf not in '七年级，八年级，九年级') or (
                cell_cd == '高中（含职业高中、技校）' and cell_cf not in '高中一年级，高中二年级，高中三年级，中职一年级，中职二年级，中职三年级') or (
                cell_cd == '大专（含高职或高专）' and cell_cf not in '高职一年级，高职二年级，高职三年级，大专') or (
                cell_cd == '本科及以上' and cell_cf not in '本科，硕士或博士研究生'):
            print('异常数据' + str(i) + '行')
            # 如果在校生状态不等于在校，或者在校生状况为-或‘特殊学校’，则比对下一行数据
            if cell_ce == '在校' and cell_cf not in '-,特殊学校':
                # 放入数组
                row1 = [i - 1, cell_b, cell_c, cell_bq, cell_bs, cell_bp, cell_bq, cell_bs, cell_bz, cell_cd, cell_ce,
                        cell_cf,
                        cell_cj,
                        cell_du]
                setSQLData1.append(row1)

        # 16-16（含）-60周岁（不含）健康男性非在校生为无劳动力或弱劳动力或半劳动力
        if 16 <= cell_bz < 60 and cell_cp == '是' and cell_bq == '男' and cell_ce != '在校' and cell_ck not in '普通劳动力，技能劳动力':
            print('异常数据，16-16（含）-60周岁（不含）健康男性非在校生为无劳动力或弱劳动力或半劳动力' + str(i) + '行')
            # 放入数组
            row1 = [i - 1, cell_b, cell_c, cell_bq, cell_bs, cell_bp, cell_bq, cell_bs, cell_bz, cell_ce,
                    cell_cp,
                    cell_ck,
                    cell_du]
            setSQLData2.append(row2)

        # 将数据写入Excel
    wb = Workbook()
    # 设置Excel文件名
    dest_filename = '分析结果数据（学生在校信息）' + time.strftime('%m%d%H%M%S', time.localtime(time.time())) + '.xlsx'

    # 新建一个表
    ws1 = wb.active

    # 填写表内容
    for listIndex in range(len(setSQLData)):
        ws1.append(setSQLData[listIndex])

    wb.save(filename=dest_filename)

except:
    print(traceback.format_exc())
    with open('log.txt', 'w') as file_object:
        file_object.write(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())) + "\n")
        file_object.write("打印异常日志：\n")
        file_object.write(traceback.format_exc())
