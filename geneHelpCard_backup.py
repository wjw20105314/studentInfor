#!/usr/bin/env python
# -*- coding: utf-8 -*-

# coding=UTF-8
import os
import traceback
from decimal import Decimal

import openpyxl
import configparser
import time
from docxtpl import DocxTemplate
import sys
import os
import win32com.client


def get_path(path):
    r = os.path.abspath(path)
    return r
def doc2pdf(src, dst):
    src = get_path(src)
    dst = get_path(dst)
    if not os.path.exists(src):
        print(src + "不存在，无法继续！")
        return False
    os.system('taskkill /im wps.exe')
    # 如果文件存在就删除
    if os.path.exists(dst):
        os.remove(dst)
    o = win32com.client.Dispatch("Kwps.Application")
    o.Visible = False
    doc = o.Documents.Open(src)
    doc.ExportAsFixedFormat(dst, 17)
    o.Quit()
    if os.path.exists(dst):
        return True
    else:
        return False

try:

    config = configparser.ConfigParser()
    config.read('properties_help.conf', encoding='utf-8-sig')

    file_name1 = config['fileCofig']['file_name1']  # 帮扶干部查询
    file_name2 = config['fileCofig']['file_name2']  # 措施年度收支统计
    file_name3 = config['fileCofig']['file_name3']  # 个人措施查询
    file_name4 = config['fileCofig']['file_name4']  # 行业部门措施查询
    file_name5 = config['fileCofig']['file_name5']  # 户措施查询
    file_name6 = config['fileCofig']['file_name6']  # 户指标查询
    file_name7 = config['fileCofig']['file_name7']  # 人口指标查询
    year = config['fileCofig']['year']
    lastyear = config['fileCofig']['lastyear']
    # 存放务工人员务工地点，key为姓名+户码，方便务工就业信息获取
    jobInfo = {}
    # 1.读取人口指标查询
    # 按每一户存储
    context7 = {}
    wb7 = openpyxl.load_workbook(file_name7)
    ws7 = wb7.active
    # 从第二行开始遍历
    for i in range(2, ws7.max_row + 1):
        # 户登记表年度为2020，则继续
        cell_7e = ws7['E' + str(i)].value
        if str(cell_7e) != str(year):
            continue
        # 获取村，户码，户主姓名
        cell_7b = ws7['B' + str(i)].value
        cell_7i = ws7['I' + str(i)].value
        cell_7f = ws7['F' + str(i)].value
        key7 = cell_7b + cell_7i + cell_7f
        # 取家庭成员
        cell_7bo = ws7['BO' + str(i)].value
        cell_7br = ws7['BR' + str(i)].value
        # 取与户主关系
        cell_7by = ws7['BY' + str(i)].value
        # 取性别
        cell_7bp = ws7['BP' + str(i)].value
        # 取年龄
        cell_7bx = int(ws7['BX' + str(i)].value)
        # 取在校生就读学校和年级
        cell_7ch = ws7['CH' + str(i)].value
        cell_7cd = ws7['CD' + str(i)].value
        # 取是否劳动力
        cell_7ci = ws7['CI' + str(i)].value
        # 取个人低保五保属性
        cell_7dg = ws7['DG' + str(i)].value
        # 取残疾类型及级别
        cell_7cq = ws7['CQ' + str(i)].value
        cell_7cr = ws7['CR' + str(i)].value
        # 取务工地点
        cell_7db = ws7['DB' + str(i)].value
        jobInfo[cell_7br] = cell_7db
        # 添加人员信息
        content_7p = {}
        # 放入户主
        person_7arr = []
        hu_7dict = {}
        hu_7dict['a1'] = cell_7bo
        hu_7dict['a2'] = cell_7by
        hu_7dict['a3'] = cell_7bp
        hu_7dict['a4'] = cell_7bx
        hu_7dict['a5'] = cell_7ch + cell_7cd
        hu_7dict['a6'] = cell_7ci
        hu_7dict['a7'] = cell_7dg
        hu_7dict['a8'] = cell_7cq + cell_7cr
        # 判断key是否存在，不存在则新加
        if key7 not in context7.keys():

            # 家庭成员为户主
            if cell_7f == cell_7bo:
                # 直接作为户主放入
                person_7arr.append(hu_7dict)
            else:
                # 将户主放前面，再放普通成员
                first_7dict = {'a1': cell_7f}
                person_7arr.append(first_7dict)
                person_7arr.append(hu_7dict)

            # 将户加入字典中
            content_7p['more1'] = person_7arr
            yanglren = 0
            #60岁买养老保险人数
            if cell_7bx >=60:
                yanglren = 1
            content_7p['yanglren'] = yanglren
            context7[key7] = content_7p
        else:
            # 存在则取出处理,同时要处理户主
            content_7p = context7[key7]
            # 60岁买养老保险人数
            if cell_7bx >= 60:
                yanglren = content_7p['yanglren']
                yanglren = yanglren + 1
                content_7p['yanglren'] = yanglren
            person_7arr = content_7p['more1']
            # 取出户主姓名，第一个
            first_7dict = person_7arr[0]
            a1 = first_7dict['a1']
            if a1 == cell_7bo:
                person_7arr[0] = hu_7dict  # 更新
            else:
                person_7arr.append(hu_7dict)

    # 2.读取户指标查询
    # 按每一户存储
    context6 = {}
    wb6 = openpyxl.load_workbook(file_name6)
    ws6 = wb6.active
    # 从第二行开始遍历
    for i in range(2, ws6.max_row + 1):
        # 户登记表年度为2020，则继续
        cell_6f = ws6['F' + str(i)].value
        if str(cell_6f) != str(year):
            continue
        # 获取村，户码，户主姓名
        cell_6c = ws6['C' + str(i)].value
        cell_6h = ws6['H' + str(i)].value
        cell_6g = ws6['G' + str(i)].value
        key6 = cell_6c + cell_6h + cell_6g
        # 获取村名，截取倒数第三位以前
        cun = cell_6c[0:-3]
        # 取致贫原因
        cell_6ah = ws6['AH' + str(i)].value
        cell_6ai = ws6['AI' + str(i)].value
        # 取贫困户属性
        cell_6ad = ws6['AD' + str(i)].value
        # 取劳动力数
        cell_6w = ws6['W' + str(i)].value
        # 取脱贫年度,取前四位
        cell_6cf = ws6['CF' + str(i)].value
        tuop = cell_6cf[0:4]
        # 危房改造
        cell_6bx = ws6['BX' + str(i)].value
        cell_6bw = ws6['BW' + str(i)].value

        # 放入户主
        hu_6dict = {}
        hu_6dict['cun'] = cun
        hu_6dict['hu'] = cell_6g
        hu_6dict['reson'] = cell_6ah + ',' + cell_6ai
        hu_6dict['shux'] = cell_6ad
        hu_6dict['laod'] = cell_6w
        hu_6dict['tuop'] = tuop
        hu_6dict['weifn'] = cell_6bx
        hu_6dict['weifd'] = cell_6bw
        context6[key6] = hu_6dict

    # 3.读取个人措施查询
    # 按每一户存储
    context3 = {}
    wb3 = openpyxl.load_workbook(file_name3)
    ws3 = wb3.active
    # 从第二行开始遍历
    for i in range(2, ws3.max_row + 1):
        # 户登记表年度为2020，则继续
        cell_3o = ws3['O' + str(i)].value
        if str(cell_3o) != str(year):
            continue
        # 获取村、户码、户主姓名
        cell_3b = ws3['B' + str(i)].value
        cell_3c = ws3['C' + str(i)].value
        cell_3d = ws3['D' + str(i)].value

        key3 = cell_3b + cell_3c + cell_3d
        # 取成员姓名,身份证
        cell_3l = ws3['L' + str(i)].value
        cell_3m = ws3['M' + str(i)].value
        # 取项目类型
        cell_3p = ws3['P' + str(i)].value
        # 取实际收益，养老
        cell_3w = float(ws3['W' + str(i)].value)
        # 取实际投入，教育，养老
        cell_3y = float(ws3['Y' + str(i)].value)
        # 添加人员信息
        # 判断key是否存在，不存在则新加
        if key3 not in context3.keys():
            first_3dict = {}
            # 教育保障
            if cell_3p == '教育扶贫':
                # 放入人数和教育生活补助
                first_3dict['jyren'] = 1
                first_3dict['jybu'] = cell_3y
            elif cell_3p == '养老保险':
                # 放入养老投入和收益
                first_3dict['yangr'] = 1
                first_3dict['yangzy'] = cell_3y
                first_3dict['yangn'] = cell_3w
            elif cell_3p == '医疗保险':
                # 放入投入和收益
                first_3dict['yir'] = 1
                first_3dict['yiz'] = cell_3y
            elif cell_3p == '医疗救助':
                # 放入收益
                first_3dict['yij'] = cell_3w
            elif cell_3p == '技能培训':
                # 放入人数
                first_3dict['jishir'] = 1
                first_3dict['jizyr'] = 1
            elif cell_3p == '就业扶贫':
                # 构造数组
                job_3dict = {}
                # 务工人员
                job_3dict['b1'] = cell_3l
                # 务工企业
                key_job = cell_3m
                if key_job in jobInfo.keys():
                    job_3dict['b2'] = jobInfo[key_job]
                # 年收入
                job_3dict['b3'] = cell_3w
                job_3arr = []
                job_3arr.append(job_3dict)
                first_3dict['more2'] = job_3arr
            context3[key3] = first_3dict
        else:
            # 存在则取出处理
            first_3dict = context3[key3]
            # 教育保障
            if cell_3p == '教育扶贫':
                # 放入人数和教育生活补助
                if 'jyren' in first_3dict.keys():
                    first_3dict['jyren'] = first_3dict['jyren'] + 1
                    first_3dict['jybu'] = first_3dict['jybu'] + cell_3y
                else:
                    first_3dict['jyren'] = 1
                    first_3dict['jybu'] = cell_3y

            elif cell_3p == '养老保险':
                # 放入养老投入和收益
                if 'yangr' in first_3dict.keys():
                    first_3dict['yangr'] = first_3dict['yangr'] + 1
                    first_3dict['yangzy'] = first_3dict['yangzy'] + cell_3y
                    first_3dict['yangn'] = first_3dict['yangn'] + cell_3w
                else:
                    first_3dict['yangr'] = 1
                    first_3dict['yangzy'] = cell_3y
                    first_3dict['yangn'] = cell_3w

            elif cell_3p == '医疗保险':
                # 放入投入和收益
                if 'yir' in first_3dict.keys():
                    first_3dict['yir'] = first_3dict['yir'] + 1
                    first_3dict['yiz'] = first_3dict['yiz'] + cell_3y
                else:
                    first_3dict['yir'] = 1
                    first_3dict['yiz'] = cell_3y
            elif cell_3p == '医疗救助':
                # 放入收益
                if 'yij' in first_3dict.keys():
                    first_3dict['yij'] = first_3dict['yij'] + cell_3w
                else:
                    first_3dict['yij'] = cell_3w
            elif cell_3p == '技能培训':
                # 放入人数
                if 'jishir' in first_3dict.keys():
                    first_3dict['jishir'] = first_3dict['jishir'] + 1
                    first_3dict['jizyr'] = first_3dict['jizyr'] + 1
                else:
                    first_3dict['jishir'] = 1
                    first_3dict['jizyr'] = 1
            elif cell_3p == '就业扶贫':
                job_3dict = {}
                # 务工人员
                job_3dict['b1'] = cell_3l
                # 务工企业
                key_job = cell_3m
                if key_job in jobInfo.keys():
                    job_3dict['b2'] = jobInfo[key_job]
                # 年收入
                job_3dict['b3'] = cell_3w

                if 'more2' in first_3dict.keys():
                    job_3arr = first_3dict['more2']
                    job_3arr.append(job_3dict)
                else:
                    job_3new = []
                    job_3new.append(job_3dict)
                    first_3dict['more2'] = job_3new

    # 4.行业部门措施查询
    # 按每一户存储
    context4 = {}
    wb4 = openpyxl.load_workbook(file_name4)
    ws4 = wb4.active
    # 从第二行开始遍历
    for i in range(2, ws4.max_row + 1):
        # 户登记表年度为2020，或者上一年度12月，则继续
        cell_4p = ws4['P' + str(i)].value
        # 取月份
        cell_4q = int(ws4['Q' + str(i)].value)

        if str(cell_4p) != str(year) and (str(cell_4p) != str(lastyear) or cell_4q != 12):
            continue
        # 获取村、户码、户主姓名
        cell_4e = ws4['E' + str(i)].value
        cell_4f = ws4['F' + str(i)].value
        cell_4g = ws4['G' + str(i)].value
        key4 = cell_4e + cell_4f + cell_4g
        # 取补贴成员
        cell_4j = ws4['J' + str(i)].value
        # 取项目名称
        cell_4n = ws4['N' + str(i)].value

        # 取发放金额
        cell_4r = Decimal(float(ws4['R' + str(i)].value)).quantize(Decimal('0.00'))
        # 添加人员信息
        # 判断key是否存在，不存在则新加
        if key4 not in context4.keys():
            first_4dict = {}
            # 耕地
            if cell_4n.find('耕地') != -1:
                # 计算单项金额
                first_4dict['gendi'] = cell_4r
            elif cell_4n.find('生态林') != -1:
                first_4dict['shengt'] = cell_4r
            elif cell_4n.find('残疾人生活津贴') != -1:
                first_4dict['kuncj'] = cell_4r
            elif cell_4n.find('重度残疾人护理') != -1:
                first_4dict['zhongcj'] = cell_4r
            elif cell_4n.find('计划生育') != -1:
                first_4dict['jisheng'] = cell_4r
            elif cell_4n.find('低保金') != -1:
                # 保障人员
                first_4dict['bzry'] = cell_4j
                # 总额
                first_4dict['dbzong'] = cell_4r
                first_4dict['dizon'] = cell_4r
                # 低保每月
                first_4dict['dijin'] = cell_4r / cell_4q
                # 保存月份
                first_4dict['dbyf'] = cell_4q
                # 保障人数
                first_4dict['diren'] = 1
            elif cell_4n.find('五保金') != -1:
                # 保障人员
                first_4dict['bzry'] = cell_4j
                # 总额
                first_4dict['wbzong'] = cell_4r
                first_4dict['dizon'] = cell_4r
                # 五保每月
                first_4dict['wujin'] = cell_4r / cell_4q
                # 保存月份
                first_4dict['wbyf'] = cell_4q
                # 保障人数
                first_4dict['diren'] = 1
            else:
                # 统计其他
                first_4dict['zhengqt'] = cell_4r
            context4[key4] = first_4dict
        else:
            # 存在则取出处理
            first_4dict = context4[key4]
            # 耕地
            if cell_4n.find('耕地') != -1:
                if 'gendi' in first_4dict.keys():
                    first_4dict['gendi'] = cell_4r + first_4dict['gendi']
                else:
                    # 计算单项金额
                    first_4dict['gendi'] = cell_4r

            elif cell_4n.find('生态林') != -1:
                if 'shengt' in first_4dict.keys():
                    first_4dict['shengt'] = cell_4r + first_4dict['shengt']
                else:
                    # 计算单项金额
                    first_4dict['shengt'] = cell_4r
            elif cell_4n.find('残疾人生活津贴') != -1:
                if 'kuncj' in first_4dict.keys():
                    first_4dict['kuncj'] = cell_4r + first_4dict['kuncj']
                else:
                    first_4dict['kuncj'] = cell_4r
            elif cell_4n.find('重度残疾人护理') != -1:
                if 'zhongcj' in first_4dict.keys():
                    first_4dict['zhongcj'] = cell_4r + first_4dict['zhongcj']
                else:
                    first_4dict['zhongcj'] = cell_4r
            elif cell_4n.find('计划生育') != -1:
                if 'jisheng' in first_4dict.keys():
                    first_4dict['jisheng'] = cell_4r + first_4dict['jisheng']
                else:
                    first_4dict['jisheng'] = cell_4r
            elif cell_4n.find('低保金') != -1:
                # 比较月份
                dbyf = cell_4q
                if 'dizon' in first_4dict.keys():
                    first_4dict['dizon'] = cell_4r + first_4dict['dizon']
                else:
                    first_4dict['dizon'] = cell_4r

                if 'dbzong' in first_4dict.keys():
                    first_4dict['dbzong'] = cell_4r + first_4dict['dbzong']
                    # 保障人员
                    bzry = first_4dict['bzry']
                    if bzry != cell_4j:
                        # 有新的保障人员
                        first_4dict['diren'] = first_4dict['diren'] + 1
                        first_4dict['bzry'] = cell_4j
                    # 取大的月份
                    if cell_4q > first_4dict['dbyf']:
                        first_4dict['dbyf'] = cell_4q
                    else:
                        dbyf = first_4dict['dbyf']
                else:
                    first_4dict['dbzong'] = cell_4r
                    first_4dict['dbyf'] = cell_4q
                    first_4dict['diren'] = 1
                    # 保障人员
                    first_4dict['bzry'] = cell_4j
                # 低保每月
                first_4dict['dijin'] = first_4dict['dbzong'] / dbyf

            elif cell_4n.find('五保金') != -1:
                # 比较月份
                wbyf = cell_4q
                if 'dizon' in first_4dict.keys():
                    first_4dict['dizon'] = cell_4r + first_4dict['dizon']
                else:
                    first_4dict['dizon'] = cell_4r

                if 'wbzong' in first_4dict.keys():
                    first_4dict['wbzong'] = cell_4r + first_4dict['wbzong']
                    # 保障人员
                    bzry = first_4dict['bzry']
                    if bzry != cell_4j:
                        # 有新的保障人员
                        first_4dict['diren'] = first_4dict['diren'] + 1
                        first_4dict['bzry'] = cell_4j
                    # 取大的月份
                    if cell_4q > first_4dict['wbyf']:
                        first_4dict['wbyf'] = cell_4q
                    else:
                        wbyf = first_4dict['wbyf']
                else:
                    first_4dict['wbzong'] = cell_4r
                    first_4dict['wbyf'] = cell_4q
                    first_4dict['diren'] = 1
                    # 保障人员
                    first_4dict['bzry'] = cell_4j
                # 低保每月
                first_4dict['wujin'] = first_4dict['wbzong'] / dbyf
            else:
                # 统计其他
                if 'zhengqt' in first_4dict.keys():
                    first_4dict['zhengqt'] = cell_4r + first_4dict['zhengqt']
                else:
                    first_4dict['zhengqt'] = cell_4r

    # 5.读取户措施查询
    # 按每一户存储
    context5 = {}
    wb5 = openpyxl.load_workbook(file_name5)
    ws5 = wb5.active
    # 从第二行开始遍历
    for i in range(2, ws5.max_row + 1):
        # 户登记表年度为2020，则继续
        cell_5m = ws5['M' + str(i)].value
        # 取项目类型
        cell_5n = ws5['N' + str(i)].value
        if str(cell_5m) != str(year) or cell_5n not in ('产业扶贫', '资产扶贫'):
            continue
        # 获取村，户码，户主姓名
        cell_5b = ws5['B' + str(i)].value
        cell_5i = ws5['I' + str(i)].value
        cell_5j = ws5['J' + str(i)].value
        key5 = cell_5b + cell_5i + cell_5j

        # 取项目名称
        cell_5o = ws5['O' + str(i)].value
        # 取实际收益
        cell_5u = float(ws5['U' + str(i)].value)
        # 取实际投入
        cell_5w = float(ws5['W' + str(i)].value)
        # 产业扶贫拼接内容
        chanye = ''
        # 添加人员信息
        hu_5dict = {}
        if cell_5n == '产业扶贫':
            chanye = cell_5n + cell_5o + '，' + '实际投入' + str(cell_5w) + '元，' + '实际收益' + str(cell_5u) + '元，' + '盈利' + str(
                cell_5u - cell_5w)+'元'
            hu_5dict['changymx'] = chanye
        elif cell_5n == '资产扶贫':
            hu_5dict['zichanfh'] = cell_5u
        # 判断key是否存在，不存在则新加
        if key5 not in context5.keys():

            context5[key5] = hu_5dict
        else:
            hu_5dict = context5[key5]
            if 'changymx' in hu_5dict.keys() and cell_5n == '产业扶贫':
                hu_5dict['changymx'] = hu_5dict['changymx'] + ';' + chanye
            elif 'zichanfh' in hu_5dict.keys() and cell_5n == '资产扶贫':
                hu_5dict['zichanfh'] = hu_5dict['zichanfh'] + cell_5u
            else:
                if cell_5n == '产业扶贫':
                    hu_5dict['changymx'] = chanye
                elif cell_5n == '资产扶贫':
                    hu_5dict['zichanfh'] = cell_5u

    # 6.读取措施年度收支统计
    # 按每一户存储
    context2 = {}
    wb2 = openpyxl.load_workbook(file_name2)
    ws2 = wb2.active
    # 从第4行开始遍历
    for i in range(4, ws2.max_row + 1):
        # 户登记表年度为2020，则继续
        cell_2o = ws2['O' + str(i)].value
        if str(cell_2o) != str(year) or cell_2o == '':
            continue
        # 获取村，户码，户主姓名
        cell_2d = ws2['D' + str(i)].value
        cell_2e = ws2['E' + str(i)].value
        cell_2f = ws2['F' + str(i)].value
        key2 = cell_2d + cell_2e + cell_2f

        # 取总收入
        cell_2p = float(ws2['P' + str(i)].value)
        # 取工资性收入
        cell_2q = float(ws2['Q' + str(i)].value)
        # 取家庭经营性收入
        cell_2r = float(ws2['R' + str(i)].value)
        # 取财产性性收入
        cell_2s = float(ws2['S' + str(i)].value)
        # 取转移性性收入
        cell_2t = float(ws2['T' + str(i)].value)
        # 医保保销
        cell_2aa = float(ws2['AA' + str(i)].value)
        # 医疗救助
        cell_2ab = float(ws2['AB' + str(i)].value)
        # 慰问金
        cell_2ac = float(ws2['AC' + str(i)].value)
        # 以奖代补
        cell_2ad = float(ws2['AD' + str(i)].value)
        # 总支出
        cell_2ae = float(ws2['AE' + str(i)].value)
        # 生成经营性支出
        cell_2af = float(ws2['AF' + str(i)].value)
        # 转移性支出
        cell_2ag = float(ws2['AG' + str(i)].value)
        # 社会保障支出
        cell_2ai = float(ws2['AI' + str(i)].value)
        # 其他转移性支出
        cell_2ak = float(ws2['AK' + str(i)].value)

        # 年度家庭可支配收入
        cell_2al = float(ws2['AL' + str(i)].value)
        # 家庭年人均可支配收入
        cell_2am = float(ws2['AM' + str(i)].value)

        # 添加人员信息
        hu_2dict = {}
        # 取总收入
        hu_2dict['jtzongs'] = cell_2p
        # 取工资性收入
        hu_2dict['gzsr'] = cell_2q
        # 取家庭经营性收入
        hu_2dict['jysr'] = cell_2r
        # 取财产性性收入
        hu_2dict['ccsr'] = cell_2s
        # 取转移性性收入
        hu_2dict['zzsr'] = cell_2t
        # 医保保销
        hu_2dict['jyibaoxiao'] = cell_2aa
        # 医疗救助
        hu_2dict['jyijiuzhu'] = cell_2ab
        # 慰问金
        hu_2dict['jweijin'] = cell_2ac
        # 以奖代补
        hu_2dict['jiajiangbu'] = cell_2ad
        # 总支出
        hu_2dict['jtzongz'] = cell_2ae
        # 生成经营性支出
        hu_2dict['jyzc'] = cell_2af
        # 转移性支出
        hu_2dict['zyzc'] = cell_2ag
        # 社会保障支出
        hu_2dict['jiasb'] = cell_2ai
        # 其他转移性支出
        hu_2dict['jiaqizhichu'] = cell_2ak

        # 年度家庭可支配收入
        hu_2dict['jtzhipeisr'] = cell_2al
        # 家庭年人均可支配收入
        hu_2dict['jtrensr'] = cell_2am
        context2[key2] = hu_2dict

    # 7.读取帮扶干部查询
    # 按每一户存储
    context1 = {}
    wb1 = openpyxl.load_workbook(file_name1)
    ws1 = wb1.active
    # 从第2行开始遍历
    for i in range(2, ws1.max_row + 1):
        # 户登记表年度为2020，帮扶状态为正在帮扶，则继续
        cell_1d = ws1['D' + str(i)].value
        cell_1q = ws1['Q' + str(i)].value
        if str(cell_1d) != str(year) or cell_1q != '正在帮扶':
            continue
        # 获取村，户码，户主姓名
        cell_1b = ws1['B' + str(i)].value
        cell_1g = ws1['G' + str(i)].value
        cell_1f = ws1['F' + str(i)].value
        key1 = cell_1b + cell_1g + cell_1f

        # 取帮扶单位
        cell_1ac = ws1['AC' + str(i)].value
        # 取帮扶干部
        cell_1r = ws1['R' + str(i)].value
        # 取联系电话
        cell_1u = ws1['U' + str(i)].value

        # 添加人员信息
        hu_1dict = {}
        # 取帮扶单位
        hu_1dict['bfdw'] = cell_1ac
        # 取帮扶干部
        hu_1dict['bfgb'] = cell_1r
        # 取联系电话
        hu_1dict['lxdh'] = cell_1u
        context1[key1] = hu_1dict

    # 循环# 2.读取户指标查询，取出每一户，并合并其他数据
    for key in context2.keys():
        dict_2 = context2[key]
        if key in context1.keys():
            dict_1 = context1[key]
            dict_2.update(dict_1)

        if key in context3.keys():
            dict_3 = context3[key]
            dict_2.update(dict_3)

        if key in context4.keys():
            dict_4 = context4[key]
            dict_2.update(dict_4)

        if key in context5.keys():
            dict_5 = context5[key]
            dict_2.update(dict_5)

        if key in context6.keys():
            dict_6 = context6[key]
            dict_2.update(dict_6)

        if key in context7.keys():
            dict_7 = context7[key]
            dict_2.update(dict_7)

        # 取出每一户的信息，生成一份word
        word = dict_2
        doc = DocxTemplate("model.docx")  # 模板文档
        cunName = dict_2['cun']
        path = ".\\" + cunName + "村帮扶明白卡\\"
        if not os.path.exists(path):
            os.makedirs(path)

        doc.render(word)  # 执行替换

        fileName = path + key + ".docx"
        output = path + key + ".pdf"
        print(fileName)
        doc.save(fileName)
        doc2pdf(fileName, output)
        os.remove(fileName)




except:
    print(traceback.format_exc())
    with open('log.txt', 'w') as file_object:
        file_object.write(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())) + "\n")
        file_object.write("打印异常日志：\n")
        file_object.write(traceback.format_exc())



