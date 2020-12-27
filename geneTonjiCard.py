#!/usr/bin/env python
# -*- coding: utf-8 -*-

# coding=UTF-8

import traceback

import openpyxl
import configparser
import time
from docxtpl import DocxTemplate

import os

try:

    config = configparser.ConfigParser()
    config.read('properties_tonji.conf', encoding='utf-8-sig')

    file_name1 = config['fileCofig']['file_name1']  # 措施年度收支统计18
    file_name2 = config['fileCofig']['file_name2']  # 措施年度收支统计19
    file_name3 = config['fileCofig']['file_name3']  # 措施年度收支统计20

    # 1.读取措施年度收支统计-18
    # 按每一户存储
    context1 = {}
    wb1 = openpyxl.load_workbook(file_name1)
    ws1 = wb1.active
    # 从第4行开始遍历
    for i in range(4, ws1.max_row + 1):
        # 最后一行有空格
        cell_2o_18 = ws1['O' + str(i)].value
        if cell_2o_18 == '' or cell_2o_18 is None:
            continue
        # 取市、县、镇，倒数最后一个字不取
        cell_2a_18 = ws1['a' + str(i)].value
        shi = cell_2a_18[0:-1]
        cell_2b_18 = ws1['b' + str(i)].value
        xian = cell_2b_18[0:-1]
        cell_2c_18 = ws1['c' + str(i)].value
        zhen = cell_2c_18[0:-1]

        # 获取村，户码，户主姓名
        cell_2d_18 = ws1['D' + str(i)].value
        cun = cell_2d_18[0:-3]
        cell_2e_18 = ws1['E' + str(i)].value
        cell_2f_18 = ws1['F' + str(i)].value
        key1 = cell_2d_18 + cell_2e_18 + cell_2f_18

        # 年末人口
        cell_2k_18 = int(ws1['K' + str(i)].value)

        # 取总收入
        cell_2p_18 = float(ws1['P' + str(i)].value)
        # 取工资性收入
        cell_2q_18 = float(ws1['Q' + str(i)].value)
        # 取家庭经营性收入
        cell_2r_18 = float(ws1['R' + str(i)].value)
        # 取转移性性收入
        cell_2t_18 = float(ws1['T' + str(i)].value)
        # 取财产性性收入
        cell_2s_18 = float(ws1['S' + str(i)].value)

        # 生成经营性支出
        cell_2af_18 = float(ws1['AF' + str(i)].value)
        # 转移性支出
        cell_2ag_18 = float(ws1['AG' + str(i)].value)
        # 家庭年人均可支配收入
        cell_2am_18 = float(ws1['AM' + str(i)].value)

        # 添加人员信息
        hu_1dict = {}
        # 市
        hu_1dict['shi'] = shi
        # 县
        hu_1dict['xian'] = xian
        # 镇
        hu_1dict['zhen'] = zhen
        # 村
        hu_1dict['cun'] = cun
        # 取总人口
        hu_1dict['ren18'] = cell_2k_18
        # 取总收入
        hu_1dict['zs18'] = cell_2p_18
        # 取工资性收入
        hu_1dict['gz18'] = cell_2q_18
        # 取家庭经营性收入
        hu_1dict['jysr18'] = cell_2r_18
        # 取转移性性收入
        hu_1dict['zysr18'] = cell_2t_18
        # 取财产性性收入
        hu_1dict['ccsr18'] = cell_2s_18
        # 生成经营性支出
        hu_1dict['jyzc18'] = cell_2af_18
        # 转移性支出
        hu_1dict['zyzc18'] = cell_2ag_18
        # 家庭年人均可支配收入
        hu_1dict['kzp18'] = cell_2am_18

        context1[key1] = hu_1dict

    # 2.读取措施年度收支统计-19
    # 按每一户存储
    context2 = {}
    wb2 = openpyxl.load_workbook(file_name2)
    ws2 = wb2.active
    # 从第4行开始遍历
    for i in range(4, ws2.max_row + 1):
        # 最后一行有空格
        cell_2o_19 = ws2['O' + str(i)].value
        if cell_2o_19 == '' or cell_2o_19 is None:
            continue

        # 取市、县、镇，倒数最后一个字不取
        cell_2a_19 = ws2['a' + str(i)].value
        shi = cell_2a_19[0:-1]
        cell_2b_19 = ws2['b' + str(i)].value
        xian = cell_2b_19[0:-1]
        cell_2c_19 = ws2['c' + str(i)].value
        zhen = cell_2c_19[0:-1]

        # 获取村，户码，户主姓名
        cell_2d_19 = ws2['D' + str(i)].value
        cun = cell_2d_19[0:-3]
        cell_2e_19 = ws2['E' + str(i)].value
        cell_2f_19 = ws2['F' + str(i)].value
        key2 = cell_2d_19 + cell_2e_19 + cell_2f_19

        # 年末人口
        cell_2k_19 = int(ws2['K' + str(i)].value)

        # 取总收入
        cell_2p_19 = float(ws2['P' + str(i)].value)
        # 取工资性收入
        cell_2q_19 = float(ws2['Q' + str(i)].value)
        # 取家庭经营性收入
        cell_2r_19 = float(ws2['R' + str(i)].value)
        # 取转移性性收入
        cell_2t_19 = float(ws2['T' + str(i)].value)
        # 取财产性性收入
        cell_2s_19 = float(ws2['S' + str(i)].value)

        # 生成经营性支出
        cell_2af_19 = float(ws2['AF' + str(i)].value)
        # 转移性支出
        cell_2ag_19 = float(ws2['AG' + str(i)].value)
        # 家庭年人均可支配收入
        cell_2am_19 = float(ws2['AM' + str(i)].value)

        # 添加人员信息
        hu_2dict = {}
        # 市
        hu_2dict['shi'] = shi
        # 县
        hu_2dict['xian'] = xian
        # 镇
        hu_2dict['zhen'] = zhen
        # 村
        hu_2dict['cun'] = cun
        # 取总人口
        hu_2dict['ren19'] = cell_2k_19
        # 取总收入
        hu_2dict['zs19'] = cell_2p_19
        # 取工资性收入
        hu_2dict['gz19'] = cell_2q_19
        # 取家庭经营性收入
        hu_2dict['jysr19'] = cell_2r_19
        # 取转移性性收入
        hu_2dict['zysr19'] = cell_2t_19
        # 取财产性性收入
        hu_2dict['ccsr19'] = cell_2s_19
        # 生成经营性支出
        hu_2dict['jyzc19'] = cell_2af_19
        # 转移性支出
        hu_2dict['zyzc19'] = cell_2ag_19
        # 家庭年人均可支配收入
        hu_2dict['kzp19'] = cell_2am_19

        context2[key2] = hu_2dict

    # 3.读取措施年度收支统计-20
    # 按每一户存储
    context3 = {}
    wb3 = openpyxl.load_workbook(file_name3)
    ws3 = wb3.active
    # 从第4行开始遍历
    for i in range(4, ws3.max_row + 1):
        # 最后一行有空格
        cell_2o_20 = ws3['O' + str(i)].value
        if cell_2o_20 == '' or cell_2o_20 is None:
            continue
        # 取市、县、镇，倒数最后一个字不取
        cell_2a_20 = ws3['a' + str(i)].value
        shi = cell_2a_20[0:-1]
        cell_2b_20 = ws3['b' + str(i)].value
        xian = cell_2b_20[0:-1]
        cell_2c_20 = ws3['c' + str(i)].value
        zhen = cell_2c_20[0:-1]

        # 获取村，户码，户主姓名
        cell_2d_20 = ws3['D' + str(i)].value
        cun = cell_2d_20[0:-3]
        cell_2e_20 = ws3['E' + str(i)].value
        cell_2f_20 = ws3['F' + str(i)].value
        key3 = cell_2d_20 + cell_2e_20 + cell_2f_20

        # 年末人口
        cell_2k_20 = int(ws3['K' + str(i)].value)

        # 取总收入
        cell_2p_20 = float(ws3['P' + str(i)].value)
        # 取工资性收入
        cell_2q_20 = float(ws3['Q' + str(i)].value)
        # 取家庭经营性收入
        cell_2r_20 = float(ws3['R' + str(i)].value)
        # 取转移性性收入
        cell_2t_20 = float(ws3['T' + str(i)].value)
        # 取财产性性收入
        cell_2s_20 = float(ws3['S' + str(i)].value)

        # 生成经营性支出
        cell_2af_20 = float(ws3['AF' + str(i)].value)
        # 转移性支出
        cell_2ag_20 = float(ws3['AG' + str(i)].value)
        # 家庭年人均可支配收入
        cell_2am_20 = float(ws3['AM' + str(i)].value)

        # 添加人员信息
        hu_3dict = {}
        # 市
        hu_3dict['shi'] = shi
        # 县
        hu_3dict['xian'] = xian
        # 镇
        hu_3dict['zhen'] = zhen
        # 村
        hu_3dict['cun'] = cun

        #户主姓名
        hu_3dict['xingming'] = cell_2f_20
        # 户属性
        cell_2j_20 = ws3['J' + str(i)].value
        hu_3dict['shuxing'] = cell_2j_20
        # 身份证
        cell_2g_20 = ws3['G' + str(i)].value
        hu_3dict['idcard'] = cell_2g_20

        # 取总人口
        hu_3dict['ren20'] = cell_2k_20
        # 取总收入
        hu_3dict['zs20'] = cell_2p_20
        # 取工资性收入
        hu_3dict['gz20'] = cell_2q_20
        # 取家庭经营性收入
        hu_3dict['jysr20'] = cell_2r_20
        # 取转移性性收入
        hu_3dict['zysr20'] = cell_2t_20
        # 取财产性性收入
        hu_3dict['ccsr20'] = cell_2s_20
        # 生成经营性支出
        hu_3dict['jyzc20'] = cell_2af_20
        # 转移性支出
        hu_3dict['zyzc20'] = cell_2ag_20
        # 家庭年人均可支配收入
        hu_3dict['kzp20'] = cell_2am_20

        context3[key3] = hu_3dict

    # 循环# 2.读取户指标查询，取出每一户，并合并其他数据
    for key in context3.keys():
        dict_3 = context3[key]
        if key in context1.keys():
            dict_1 = context1[key]
            dict_3.update(dict_1)

        if key in context2.keys():
            dict_2 = context2[key]
            dict_3.update(dict_2)

        # 取出每一户的信息，生成一份word
        word = dict_3
        doc = DocxTemplate("model_tongji.docx")  # 模板文档
        cunName = dict_3['cun']
        path = ".\\" + cunName + "经济收入调查表\\"
        if not os.path.exists(path):
            os.makedirs(path)

        doc.render(word)  # 执行替换

        fileName = path + key + ".docx"
        print('正在生成：'+fileName)
        doc.save(fileName)




except:
    print(traceback.format_exc())
    with open('log.txt', 'w') as file_object:
        file_object.write(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())) + "\n")
        file_object.write("打印异常日志：\n")
        file_object.write(traceback.format_exc())
