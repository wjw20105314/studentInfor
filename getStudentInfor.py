#!/usr/bin/env python
# -*- coding: utf-8 -*-

# coding=UTF-8
import traceback

import openpyxl
import configparser
import time
# 写入Excel表需要使用的库
from openpyxl import Workbook

try:

    config = configparser.ConfigParser()
    config.read('properties.conf', encoding='utf-8-sig')

    file_name = config['fileCofig']['file_name']  # Excel文件名称
    setSQLData = []
    row = ['序号', '镇', '行政村', '户主姓名', '户主证件号码', '姓名', '性别', '身份证号码', '年龄', '文化程度', '在校生状态', '在校生状况', '学校名称', '成员备注']
    setSQLData.append(row)
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    i = 1
    # print a column
    for i in range(ws.max_row):

        i = i + 1  # 下标指向下一行，第二行为数据行

        cell_k = ws['K' + str(i)].value  # 读取在校生状态
        cell_l = ws['L' + str(i)].value  # 读取在校生状况
        # 如果在校生状态不等于在校，或者在校生状况为-，则比对下一行数据
        if cell_k != '在校' or cell_l == '-':
            continue
        cell_j = ws['J' + str(i)].value  # 读取文化程度

        # 抓异常数据
        if (cell_j == '小学' and cell_l != '小学') or (cell_j == '初中' and cell_l not in '七年级，八年级，九年级') or (
                cell_j == '高中（含职业高中、技校）' and cell_l not in '高中一年级，高中二年级，高中三年级，中职一年级，中职二年级，中职三年级') or (
                cell_j == '大专（含高职或高专）' and cell_l not in '高职一年级，高职二年级，高职三年级，大专') or (
                cell_j == '本科及以上' and cell_l not in '本科，硕士或博士研究生'):
            print('异常数据' + str(i) + '行')
            cell_b = ws['B' + str(i)].value  # 读取镇
            cell_c = ws['C' + str(i)].value  # 行政村
            cell_d = ws['D' + str(i)].value  # 户主姓名
            cell_e = ws['E' + str(i)].value  # 户主证件号码
            cell_f = ws['F' + str(i)].value  # 姓名
            cell_g = ws['G' + str(i)].value  # 性别
            cell_h = ws['H' + str(i)].value  # 身份证号
            cell_i = ws['I' + str(i)].value  # 年龄
            cell_m = ws['M' + str(i)].value  # 学校名称
            cell_n = ws['N' + str(i)].value  # 成员备注
            # 放入数组
            row = [i - 1, cell_b, cell_c, cell_d, cell_e, cell_f, cell_g, cell_h, cell_i, cell_j, cell_k, cell_l,
                   cell_m,
                   cell_n]
            setSQLData.append(row)

        # 将数据写入Excel
    wb = Workbook()
    # 设置Excel文件名
    dest_filename = '分析结果数据（学生在校信息）' + time.strftime('%m%d%H%M%S', time.localtime(time.time())) + '.xlsx'

    # 新建一个表
    ws1 = wb.active

    # 填写表内容
    for listIndex in range(len(setSQLData)):
        ws1.append(setSQLData[listIndex])

    wb.save(filename=dest_filename)

except:
    print(traceback.format_exc())
    with open('log.txt', 'w') as file_object:
        file_object.write(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())) + "\n")
        file_object.write("打印异常日志：\n")
        file_object.write(traceback.format_exc())
